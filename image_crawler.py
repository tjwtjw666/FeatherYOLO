import os
import re
import requests
import openpyxl
from ultralytics import YOLO
import shutil

def get_images_from_search(keyword, page_num, save_dir, excel_file, model):
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0'
    }
    
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    
    wb_info = openpyxl.load_workbook(excel_file)
    sheet_info = wb_info['Images Info']
    
    last_row = sheet_info.max_row
    
    url = 'https://image.baidu.com/search/acjson?'
    n = 0
    
    for pn in range(0, 30 * page_num, 30):
        param = {
            'tn': 'resultjson_com',
            'logid': '7603311155072595725',
            'ipn': 'rj',
            'ct': 201326592,
            'is': '',
            'fp': 'result',
            'queryWord': keyword,
            'cl': 2,
            'lm': -1,
            'ie': 'utf-8',
            'oe': 'utf-8',
            'adpicid': '',
            'st': -1,
            'z': '',
            'ic': '',
            'hd': '',
            'latest': '',
            'copyright': '',
            'word': keyword,
            's': '',
            'se': '',
            'tab': '',
            'width': '',
            'height': '',
            'face': 0,
            'istype': 2,
            'qc': '',
            'nc': '1',
            'fr': '',
            'expermode': '',
            'force': '',
            'cg': '',
            'pn': pn,
            'rn': '30',
            'gsm': '1e',
            '1618827096642': ''
        }
        response = requests.get(url=url, headers=header, params=param)
        if response.status_code == 200:
            print('Request success.')
        response.encoding = 'utf-8'
        
        html = response.text
        image_url_list = re.findall('"thumbURL":"(.*?)",', html, re.S)
        bdImgnewsDate_list = re.findall('"bdImgnewsDate":"(.*?)",', html, re.S)
        fromPageTitle_list = re.findall('"fromPageTitle":"(.*?)",', html, re.S)
        
        for i, image_url in enumerate(image_url_list):
            try:
                image_data = requests.get(url=image_url, headers=header).content
                image_name = f'{n:06d}.jpg'
                image_path = os.path.join(save_dir, image_name)
                with open(image_path, 'wb') as fp:
                    fp.write(image_data)
                
                results = model(image_path)
                
                if results and len(results[0].boxes) > 0:
                    tail_type = results[0].names[int(results[0].boxes.cls[0])]
                else:
                    tail_type = "Unidentified"
                
                category_dir = os.path.join(save_dir, tail_type)
                if not os.path.exists(category_dir):
                    os.makedirs(category_dir)
                shutil.move(image_path, os.path.join(category_dir, image_name))
                
                bdImgnewsDate = bdImgnewsDate_list[i] if i < len(bdImgnewsDate_list) else 'N/A'
                fromPageTitle = fromPageTitle_list[i] if i < len(fromPageTitle_list) else 'N/A'
                
                print(f'Image: {image_name}, Date: {bdImgnewsDate}, Title: {fromPageTitle}, Category: {tail_type}')
                
                last_row += 1
                sheet_info.cell(row=last_row, column=1, value=image_name)
                sheet_info.cell(row=last_row, column=2, value=bdImgnewsDate)
                sheet_info.cell(row=last_row, column=3, value=fromPageTitle)
                sheet_info.cell(row=last_row, column=4, value=tail_type)
                n += 1
                wb_info.save(excel_file)
            except Exception as e:
                print(f"Failed to process image {image_name}: {e}")
                last_row += 1
                sheet_info.cell(row=last_row, column=1, value=image_name)
                sheet_info.cell(row=last_row, column=2, value="N/A")
                sheet_info.cell(row=last_row, column=3, value="N/A")
                sheet_info.cell(row=last_row, column=4, value="Failed")
                wb_info.save(excel_file)
    wb_info.close()

if __name__ == "__main__":
    model_path = "FeatherYOLO.pt"
    model = YOLO(model_path)
    
    keywords_file = 'keywords.xlsx'
    workbook = openpyxl.load_workbook(keywords_file)
    sheet = workbook['Keywords']
    keywords = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]   
    
    excel_file = './image_info.xlsx'
    
    if not os.path.exists(excel_file):
        wb_info = openpyxl.Workbook()
        sheet_info = wb_info.active
        sheet_info.title = "Images Info"
        sheet_info.append(["Image Name", "Date", "Title", "Category"])
        wb_info.save(excel_file)
    
    for keyword in keywords:
        page_num = 10
        save_dir = f'./images/{keyword}'
        get_images_from_search(keyword, page_num, save_dir, excel_file, model)